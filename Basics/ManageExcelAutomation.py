# Code to automate fixing a column value in an Excel spreadsheet and appending it in another column in the spreadsheet
# Openpyxl package is used for this purpose
# Generate a graph of the corresponding data and compare values
import openpyxl as excel
from openpyxl.chart import Reference, BarChart


def process_workbook(file_name):
    # Storing the workbook provided by the user in a custom object and utilising it later
    wb = excel.load_workbook(file_name)
    sheet = wb.active

    # Creating a new column to store values of corrected prices
    corrected_price_cell = sheet.cell(1, sheet.max_column + 1)
    corrected_price_cell.value = "corrected price"

    # Store the corrected values for price in the new column
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, sheet.max_column)
        corrected_price_cell.value = corrected_price

    # Create a chart to display the comparison between prices using the Reference and BarChart classes
    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=sheet.max_column,
                       max_col=sheet.max_column)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    # Another method to use the reference and barchart classes instead of using the 2nd import statement
    # as we are already importing the whole package at once in the first import statement
    """values = excel.chart.Reference(sheet,
                                   min_row=2,
                                   max_row=sheet.max_row,
                                   min_col=sheet.max_column,
                                   max_col=sheet.max_column)
    chart = excel.chart.BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')"""

    # Save the created workbook by overriding the existing file
    wb.save(file_name)
